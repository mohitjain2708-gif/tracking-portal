from openpyxl import load_workbook


def load_workbook_and_oonc(file_path: str):
    wb = load_workbook(file_path)
    if "OONC" not in wb.sheetnames:
        raise ValueError("OONC sheet not found")
    ws = wb["OONC"]
    return wb, ws


def detect_header_row(ws):
    keywords = ["party", "bl", "payment", "container", "train", "location", "gate", "port"]
    best_row = 1
    best_score = -1
    max_scan = min(ws.max_row, 25)

    for r in range(1, max_scan + 1):
        score = 0
        for c in range(1, ws.max_column + 1):
            txt = str(ws.cell(r, c).value or "").strip().lower()
            if any(k in txt for k in keywords):
                score += 1
        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def header_map(ws, header_row: int):
    result = {}
    for c in range(1, ws.max_column + 1):
        txt = str(ws.cell(header_row, c).value or "").strip()
        if txt:
            result[txt.lower()] = c
    return result