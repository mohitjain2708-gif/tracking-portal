import json
import os
import tempfile
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

import gspread
from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from google.oauth2.service_account import Credentials
from openpyxl import Workbook

from services.excel_parser import load_workbook_and_oonc, detect_header_row, header_map
from services.ldb_client import fetch_ldb
from services.concor_client import fetch_concor
from services.oonc_updater import update_oonc_sheet


app = FastAPI(title="Tracking Portal API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================
# PERFORMANCE SETTINGS
# =========================
MAX_WORKERS = int(os.getenv("MAX_WORKERS", "10"))
CACHE_TTL_SECONDS = int(os.getenv("CACHE_TTL_SECONDS", "1800"))  # 30 min

TRACKING_CACHE = {}
CACHE_LOCK = Lock()


def get_cached_tracking(container_no: str):
    now = time.time()
    with CACHE_LOCK:
        entry = TRACKING_CACHE.get(container_no)
        if not entry:
            return None
        if now - entry["ts"] > CACHE_TTL_SECONDS:
            del TRACKING_CACHE[container_no]
            return None
        return entry["data"]


def set_cached_tracking(container_no: str, data: dict):
    with CACHE_LOCK:
        TRACKING_CACHE[container_no] = {
            "ts": time.time(),
            "data": data,
        }


def get_google_client():
    creds_json = os.getenv("GOOGLE_CREDENTIALS_JSON", "").strip()
    if not creds_json:
        raise ValueError("GOOGLE_CREDENTIALS_JSON environment variable is missing")

    creds_dict = json.loads(creds_json)
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    return gspread.authorize(credentials)


def get_google_sheet():
    spreadsheet_id = os.getenv("GOOGLE_SHEET_ID", "").strip()
    if not spreadsheet_id:
        raise ValueError("GOOGLE_SHEET_ID environment variable is missing")

    client = get_google_client()
    return client.open_by_key(spreadsheet_id)


def get_or_create_worksheet(spreadsheet, title: str, rows: int = 1000, cols: int = 50):
    try:
        return spreadsheet.worksheet(title)
    except gspread.WorksheetNotFound:
        return spreadsheet.add_worksheet(title=title, rows=rows, cols=cols)


def load_input_sheet_as_workbook():
    input_tab = os.getenv("GOOGLE_SHEETS_INPUT_TAB", "OONC_INPUT").strip() or "OONC_INPUT"
    spreadsheet = get_google_sheet()
    ws_g = spreadsheet.worksheet(input_tab)

    values = ws_g.get_all_values()
    if not values:
        raise ValueError(f"Input sheet '{input_tab}' is empty")

    wb = Workbook()
    ws = wb.active
    ws.title = "OONC"

    for row in values:
        ws.append(row)

    return spreadsheet, wb, ws


def write_rows_to_output_sheet(spreadsheet, headers, rows):
    output_tab = os.getenv("GOOGLE_SHEETS_OUTPUT_TAB", "OONC_RESULT").strip() or "OONC_RESULT"
    ws_out = get_or_create_worksheet(
        spreadsheet,
        output_tab,
        rows=max(len(rows) + 10, 1000),
        cols=max(len(headers) + 5, 50),
    )

    data = [headers]
    for row in rows:
        data.append([row.get(h, "") for h in headers])

    ws_out.clear()
    ws_out.update("A1", data, value_input_option="USER_ENTERED")


def fetch_container_data(container_no: str) -> tuple[str, dict]:
    cached = get_cached_tracking(container_no)
    if cached is not None:
        return container_no, cached

    ldb = {}
    concor = {}
    errors = []

    try:
        ldb = fetch_ldb(container_no)
    except Exception as e:
        errors.append(f"LDB: {e}")

    try:
        concor = fetch_concor(container_no)
    except Exception as e:
        errors.append(f"CONCOR: {e}")

    result = {**ldb, **concor, "error": " | ".join(errors)}
    set_cached_tracking(container_no, result)
    return container_no, result


def process_worksheet(ws):
    hdr_row = detect_header_row(ws)
    hmap = header_map(ws, hdr_row)

    container_col = None
    for key in ["container no", "containerno", "container", "cntr no"]:
        if key in hmap:
            container_col = hmap[key]
            break

    if not container_col:
        raise ValueError("Container number column not found in OONC")

    # Collect unique containers first
    container_numbers = []
    seen = set()

    for r in range(hdr_row + 1, ws.max_row + 1):
        container_no = str(ws.cell(r, container_col).value or "").strip().upper()
        if not container_no:
            continue
        if container_no in seen:
            continue
        seen.add(container_no)
        container_numbers.append(container_no)

    tracking_map = {}

    # Parallel processing
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(fetch_container_data, cn) for cn in container_numbers]
        for future in as_completed(futures):
            cn, data = future.result()
            tracking_map[cn] = data

    update_oonc_sheet(ws, hdr_row, hmap, tracking_map)

    headers = [str(ws.cell(hdr_row, c).value or "") for c in range(1, ws.max_column + 1)]

    preview_rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        row = {}
        is_blank = True
        for c in range(1, ws.max_column + 1):
            value = ws.cell(r, c).value
            row[headers[c - 1] or f"Column {c}"] = "" if value is None else value
            if value not in [None, ""]:
                is_blank = False
        if not is_blank:
            preview_rows.append(row)

    return {
        "headers": headers,
        "rows": preview_rows,
        "tracked_containers": len(tracking_map),
        "download_ready": True,
    }


@app.get("/api/health")
async def health():
    return {
        "status": "ok",
        "cache_size": len(TRACKING_CACHE),
        "max_workers": MAX_WORKERS,
        "cache_ttl_seconds": CACHE_TTL_SECONDS,
    }


@app.post("/api/process-tracking")
async def process_tracking(file: UploadFile = File(...)):
    try:
        suffix = os.path.splitext(file.filename)[1] or ".xlsx"

        with tempfile.TemporaryDirectory() as tmpdir:
            in_path = os.path.join(tmpdir, f"input{suffix}")

            with open(in_path, "wb") as f:
                f.write(await file.read())

            wb, ws = load_workbook_and_oonc(in_path)
            result = process_worksheet(ws)

            # Skipping wb.save(out_path) for speed since frontend already uses JSON response
            return result

    except Exception as e:
        return JSONResponse(
            {
                "error": str(e),
                "error_type": type(e).__name__,
                "traceback": traceback.format_exc(),
            },
            status_code=500,
        )


@app.post("/api/sync-google-sheet")
async def sync_google_sheet():
    try:
        spreadsheet, wb, ws = load_input_sheet_as_workbook()
        result = process_worksheet(ws)
        write_rows_to_output_sheet(spreadsheet, result["headers"], result["rows"])

        return {
            **result,
            "message": "Google Sheet synced successfully",
            "input_tab": os.getenv("GOOGLE_SHEETS_INPUT_TAB", "OONC_INPUT"),
            "output_tab": os.getenv("GOOGLE_SHEETS_OUTPUT_TAB", "OONC_RESULT"),
        }

    except Exception as e:
        return JSONResponse(
            {
                "error": str(e),
                "error_type": type(e).__name__,
                "traceback": traceback.format_exc(),
            },
            status_code=500,
        )